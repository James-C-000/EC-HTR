#!/usr/bin/env python3
"""
Simple post-processing filter for EC VLM extraction results.
"""

import json
import re
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# SIMPLE FILTERS - Remove only obvious noise
# ============================================================================

# Form boilerplate that starts a line (case-insensitive)
BOILERPLATE_STARTS = [
    r'^station\s',
    r'^province\s',
    r'^observer[,:\s]',
    r'^for the month of',
    r'^lat\.?\s*\d',
    r'^long\.?\s*\d',
    r'^height above',
    r'^sums\.?\s*$',
    r'^means\.?\s*$',
    r'^total\.?\s*$',
    r'^certified correct',
    r'^do not punch',
    r'^office use',
    r'^day\s+\d+:',
]

# Skip lines that are mostly numbers (>80% numeric)
def is_mostly_numbers(line: str) -> bool:
    if not line:
        return True
    nums = sum(1 for c in line if c.isdigit() or c in '.-+')
    alphas = sum(1 for c in line if c.isalpha())
    if alphas == 0:
        return True
    return nums / (nums + alphas) > 0.8

def is_boilerplate(line: str) -> bool:
    line_lower = line.lower().strip()
    for pattern in BOILERPLATE_STARTS:
        if re.match(pattern, line_lower):
            return True
    return False

def has_excessive_repetition(text: str) -> tuple[bool, list[str]]:
    """Check if text has VLM hallucination patterns. Returns (is_artifact, unique_lines)."""
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) < 10:
        return False, lines
    
    counts = Counter(l.lower() for l in lines)
    most_common_count = counts.most_common(1)[0][1] if counts else 0
    
    # If one line appears >50% of the time and >10 times, it's an artifact
    if most_common_count > len(lines) * 0.5 and most_common_count > 10:
        # Return unique lines only
        seen = set()
        unique = []
        for line in lines:
            if line.lower() not in seen:
                seen.add(line.lower())
                unique.append(line)
        return True, unique[:15]  # Cap at 15 unique lines
    
    return False, lines

def clean_line(line: str) -> str:
    """Light cleaning of a line."""
    line = line.strip()
    # Remove leading day numbers like "1.", "15:", "31)"
    line = re.sub(r'^\d{1,2}[\.\):\s]+', '', line)
    # Remove trailing pure numbers
    line = re.sub(r'\s+\d+\.?\d*$', '', line)
    return line.strip()

def extract_remarks(raw_text: str) -> list[str]:
    """Extract remarks with minimal filtering."""
    if not raw_text:
        return []
    
    # Handle repetition artifacts
    is_artifact, lines = has_excessive_repetition(raw_text)
    
    results = []
    for line in lines:
        line = clean_line(line)
        
        # Skip short lines
        if len(line) < 3:
            continue
        
        # Skip boilerplate
        if is_boilerplate(line):
            continue
            
        # Skip mostly-numeric lines
        if is_mostly_numbers(line):
            continue
        
        results.append(line)
    
    return results

def process_results(input_file: str, output_file: str):
    """Process VLM results and create Excel output."""
    
    records = []
    with open(input_file, 'r') as f:
        for line in f:
            if line.strip():
                records.append(json.loads(line))
    
    print(f"Loaded {len(records)} records")
    
    # Process each record
    results_with_remarks = []
    results_without = []
    
    for rec in records:
        raw_text = rec.get('extracted_text', '')
        remarks = extract_remarks(raw_text)
        
        result = {
            'filename': Path(rec.get('filepath', '')).name,
            'year': rec.get('year'),
            'month': rec.get('month'),
            'station': rec.get('station_name', ''),
            'location': rec.get('location', ''),
            'remarks': '; '.join(remarks) if remarks else '',
            'raw_text': raw_text,
            'remark_count': len(remarks),
        }
        
        if remarks:
            results_with_remarks.append(result)
        else:
            results_without.append(result)
    
    print(f"Files with remarks: {len(results_with_remarks)}")
    print(f"Files without remarks: {len(results_without)}")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Sheet 1: Extracted Remarks
    ws1 = wb.active
    ws1.title = "Extracted Remarks"
    
    headers = ['Filename', 'Year', 'Month', 'Station', 'Location', 'Remarks']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row, result in enumerate(sorted(results_with_remarks, key=lambda x: (x['year'] or 0, x['month'] or 0)), 2):
        ws1.cell(row=row, column=1, value=result['filename'])
        ws1.cell(row=row, column=2, value=result['year'])
        ws1.cell(row=row, column=3, value=result['month'])
        ws1.cell(row=row, column=4, value=result['station'])
        ws1.cell(row=row, column=5, value=result['location'])
        cell = ws1.cell(row=row, column=6, value=result['remarks'])
        cell.alignment = Alignment(wrap_text=True)
    
    # Set column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 8
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 35
    ws1.column_dimensions['F'].width = 80
    
    ws1.freeze_panes = 'A2'
    
    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
    
    total_remarks = sum(r['remark_count'] for r in results_with_remarks)
    stats = [
        ("Total files processed", len(records)),
        ("Files with remarks", len(results_with_remarks)),
        ("Files without remarks", len(results_without)),
        ("Total individual remarks", total_remarks),
        ("Extraction rate", f"{len(results_with_remarks)/len(records)*100:.1f}%"),
    ]
    for row, (metric, value) in enumerate(stats, 2):
        ws2.cell(row=row, column=1, value=metric)
        ws2.cell(row=row, column=2, value=value)
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    
    # Save
    wb.save(output_file)
    print(f"\nSaved to {output_file}")
    
    # Print sample
    print("\n=== SAMPLE REMARKS ===")
    for r in results_with_remarks[:10]:
        print(f"\n{r['filename']} ({r['year']}):")
        print(f"  {r['remarks'][:200]}")

if __name__ == '__main__':
    import sys
    input_file = sys.argv[1] if len(sys.argv) > 1 else 'results.jsonl'
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'extracted_remarks.xlsx'
    process_results(input_file, output_file)
